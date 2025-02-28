import os
import io
import openpyxl
import streamlit as st
from lxml import etree
from shapely.geometry import LineString
from bs4 import BeautifulSoup
import traceback

# -------------------------------------
# 🌟 ตกแต่ง UI ด้วย CSS & Bootstrap
# -------------------------------------
st.set_page_config(page_title="KML to Excel Converter", page_icon="📄", layout="wide")

st.markdown("""
    <style>
    body {
        background-color: #f8f9fa;
    }
    .main-title {
        color: #007bff;
        text-align: center;
        font-size: 36px;
        font-weight: bold;
        margin-bottom: 10px;
    }
    .sub-title {
        color: #6c757d;
        text-align: center;
        font-size: 18px;
        margin-bottom: 20px;
    }
    .stButton>button {
        background-color: #28a745;
        color: white;
        font-size: 16px;
        border-radius: 10px;
        padding: 12px 20px;
        transition: 0.3s;
    }
    .stButton>button:hover {
        background-color: #218838;
        transform: scale(1.05);
    }
    .stDownloadButton>button {
        background-color: #007bff;
        color: white;
        font-size: 16px;
        border-radius: 10px;
        padding: 12px 20px;
        transition: 0.3s;
    }
    .stDownloadButton>button:hover {
        background-color: #0056b3;
        transform: scale(1.05);
    }
    .error-message {
        color: #dc3545;
        padding: 10px;
        border-radius: 5px;
        background-color: #f8d7da;
        border: 1px solid #f5c6cb;
        margin: 10px 0;
    }
    .success-message {
        color: #28a745;
        padding: 10px;
        border-radius: 5px;
        background-color: #d4edda;
        border: 1px solid #c3e6cb;
        margin: 10px 0;
    }
    </style>
    """, unsafe_allow_html=True)

# -------------------------------------
# 🚀 ฟังก์ชันประมวลผล
# -------------------------------------
def extract_description_data(description_html):
    """
    Extract structured data from the HTML description.
    
    Args:
        description_html (str): HTML content from KML description
        
    Returns:
        dict: Dictionary containing extracted data
    """
    try:
        soup = BeautifulSoup(description_html, 'html.parser')
        extracted_data = {}
        rows = soup.find_all('tr')
        
        for row in rows:
            cells = row.find_all('td')
            if len(cells) >= 2:
                header = cells[0].get_text(strip=True)  
                value = cells[1].get_text(strip=True)  
                extracted_data[header] = value
                
        return extracted_data
    except Exception as e:
        st.warning(f"ไม่สามารถแยกข้อมูลจาก HTML description ได้: {str(e)}")
        return {}

def parse_coordinates(coords_text):
    """
    Parse coordinate text into list of coordinate tuples.
    
    Args:
        coords_text (str): Text containing coordinates
        
    Returns:
        list: List of coordinate tuples [(lon, lat), ...]
    """
    coords = []
    try:
        for c in coords_text.split():
            parts = c.split(',')
            if len(parts) >= 2:
                coords.append((float(parts[0]), float(parts[1])))
            else:
                st.warning(f"ข้อมูลพิกัดไม่ถูกต้อง: {c}")
    except ValueError as e:
        st.warning(f"ไม่สามารถแปลงพิกัดเป็นตัวเลขได้: {str(e)}")
    
    return coords

def load_kml_lines(uploaded_file):
    """
    Load lines from a KML file and return them as Shapely LineString objects with metadata.
    
    Args:
        uploaded_file: Streamlit uploaded file object
        
    Returns:
        list: List of dictionaries containing line data
    """
    lines = []
    
    try:
        kml_data = etree.fromstring(uploaded_file.getvalue())
        ns = {'kml': 'http://www.opengis.net/kml/2.2'}
        
        # ตรวจสอบความถูกต้องของไฟล์ KML
        if kml_data.tag != '{http://www.opengis.net/kml/2.2}kml':
            st.error(f"ไฟล์ {uploaded_file.name} ไม่ใช่ไฟล์ KML ที่ถูกต้อง")
            return []
        
        for placemark in kml_data.xpath('.//kml:Placemark', namespaces=ns):
            try:
                line_string = placemark.xpath('.//kml:LineString/kml:coordinates', namespaces=ns)
                
                if line_string and line_string[0].text:
                    coords_text = line_string[0].text.strip()
                    coords = parse_coordinates(coords_text)
                    
                    if len(coords) > 1:
                        line_geom = LineString(coords)
                        
                        # ดึงชื่อของ Placemark
                        name_elements = placemark.xpath('./kml:name/text()', namespaces=ns)
                        name = name_elements[0] if name_elements else "Unnamed"
                        
                        # ดึงคำอธิบายของ Placemark
                        description_elements = placemark.xpath('./kml:description/text()', namespaces=ns)
                        description_data = {}
                        
                        if description_elements:
                            try:
                                description_data = extract_description_data(description_elements[0])
                            except Exception as e:
                                st.warning(f"ไม่สามารถแยกข้อมูลจากคำอธิบายของ '{name}' ได้: {str(e)}")
                        
                        lines.append({
                            "Name": name,
                            "Description": description_data,
                            "Line": line_geom,
                            "Start_Coordinate": coords[0] if coords else None,
                            "End_Coordinate": coords[-1] if coords else None
                        })
                    else:
                        st.warning(f"ไม่มีพิกัดเพียงพอสำหรับการสร้าง LineString: {placemark.xpath('./kml:name/text()', namespaces=ns)[0] if placemark.xpath('./kml:name/text()', namespaces=ns) else 'Unnamed'}")
                else:
                    # ข้ามข้อมูลที่ไม่มี LineString
                    pass
                    
            except Exception as e:
                name = placemark.xpath('./kml:name/text()', namespaces=ns)
                name = name[0] if name else "Unnamed"
                st.warning(f"เกิดข้อผิดพลาดในการประมวลผล Placemark '{name}': {str(e)}")
    
    except Exception as e:
        st.error(f"ไม่สามารถโหลดไฟล์ KML ได้: {str(e)}")
        st.error(traceback.format_exc())
    
    return lines

def save_to_excel_memory(lines):
    """
    Save line data to Excel in memory.
    
    Args:
        lines (list): List of line dictionaries
        
    Returns:
        BytesIO: Excel file in memory
    """
    try:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "KML Data"
        
        # กรณีไม่มีข้อมูลเลย
        if not lines:
            sheet.append(['ไม่พบข้อมูลเส้นที่ถูกต้อง'])
            output = io.BytesIO()
            workbook.save(output)
            output.seek(0)
            return output
        
        # สร้างส่วนหัวของตาราง
        headers = ['Name', 'Start_Coordinate', 'End_Coordinate']
        
        # เพิ่มส่วนหัวจากข้อมูล Description ของเส้นแรก
        description_headers = []
        for line in lines:
            if isinstance(line.get('Description'), dict) and line['Description']:
                description_headers = list(line['Description'].keys())
                break
        
        headers.extend(description_headers)
        sheet.append(headers)
        
        # เพิ่มข้อมูลแต่ละแถว
        for line in lines:
            row = [
                line.get("Name", "N/A"),
                f"{line['Start_Coordinate'][0]},{line['Start_Coordinate'][1]}" if line.get("Start_Coordinate") else "N/A",
                f"{line['End_Coordinate'][0]},{line['End_Coordinate'][1]}" if line.get("End_Coordinate") else "N/A"
            ]
            
            # เพิ่มข้อมูล Description
            if isinstance(line.get('Description'), dict):
                for header in description_headers:
                    row.append(line['Description'].get(header, "N/A"))
            else:
                # เพิ่มช่องว่างสำหรับคอลัมน์ Description ที่ไม่มีข้อมูล
                for _ in description_headers:
                    row.append("N/A")
            
            sheet.append(row)
        
        # ปรับความกว้างของคอลัมน์ให้เหมาะสม
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) if max_length < 50 else 50
            sheet.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output
        
    except Exception as e:
        st.error(f"เกิดข้อผิดพลาดในการสร้างไฟล์ Excel: {str(e)}")
        st.error(traceback.format_exc())
        # สร้างไฟล์ Excel ว่างเพื่อไม่ให้โปรแกรมล่ม
        empty_workbook = openpyxl.Workbook()
        empty_sheet = empty_workbook.active
        empty_sheet.append(['เกิดข้อผิดพลาดในการสร้างไฟล์ Excel'])
        empty_output = io.BytesIO()
        empty_workbook.save(empty_output)
        empty_output.seek(0)
        return empty_output

# -------------------------------------
# 🎯 ส่วน UI ของ Streamlit
# -------------------------------------
st.markdown('<p class="main-title">KML to Excel Converter 🚀</p>', unsafe_allow_html=True)
st.markdown('<p class="sub-title">📄 แปลงไฟล์ KML เป็น Excel พร้อมข้อมูล Metadata</p>', unsafe_allow_html=True)

with st.expander("📌 คำแนะนำการใช้งาน", expanded=False):
    st.markdown("""
    1. อัปโหลดไฟล์ KML ที่มีข้อมูลเส้น (LineString)
    2. คลิกปุ่ม "เริ่มประมวลผล" เพื่อแปลงไฟล์
    3. ดาวน์โหลดไฟล์ Excel ที่ได้
    
    **หมายเหตุ:**
    - รองรับไฟล์หลายไฟล์ในครั้งเดียว
    - แปลงเฉพาะข้อมูลประเภทเส้น (LineString) เท่านั้น
    - ดึงข้อมูลจากฟิลด์ Description ที่อยู่ในรูปแบบตาราง HTML
    """)

uploaded_files = st.file_uploader("📂 อัปโหลดไฟล์ KML", type="kml", accept_multiple_files=True)

if uploaded_files:
    # แสดงจำนวนไฟล์ที่อัปโหลด
    st.info(f"อัปโหลดไฟล์ KML จำนวน {len(uploaded_files)} ไฟล์")
    
    # ปุ่มเริ่มประมวลผล
    if st.button('⚡ เริ่มประมวลผล'):
        with st.status("⏳ กำลังประมวลผล...", expanded=True) as status:
            progress_bar = st.progress(0)
            total_files = len(uploaded_files)
            success_count = 0
            error_count = 0
            
            for i, uploaded_file in enumerate(uploaded_files):
                try:
                    st.write(f"กำลังประมวลผลไฟล์ {uploaded_file.name}...")
                    
                    # โหลดข้อมูลจากไฟล์ KML
                    lines = load_kml_lines(uploaded_file)
                    
                    if lines:
                        st.success(f"พบข้อมูลเส้นทั้งหมด {len(lines)} เส้น")
                        
                        # บันทึกข้อมูลลง Excel
                        output_memory = save_to_excel_memory(lines)
                        
                        # สร้างปุ่มดาวน์โหลด
                        st.download_button(
                            label=f"📥 ดาวน์โหลด {uploaded_file.name.replace('.kml', '.xlsx')}",
                            data=output_memory,
                            file_name=f"{uploaded_file.name.replace('.kml', '.xlsx')}",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                        success_count += 1
                    else:
                        st.warning(f"ไม่พบข้อมูลเส้นในไฟล์ {uploaded_file.name}")
                        error_count += 1
                        
                except Exception as e:
                    st.error(f"เกิดข้อผิดพลาดในการประมวลผลไฟล์ {uploaded_file.name}: {str(e)}")
                    error_count += 1
                
                # อัปเดตแถบความคืบหน้า
                progress_bar.progress((i + 1) / total_files)
            
            # สรุปผลการประมวลผล
            if success_count > 0:
                st.markdown(f'<p class="success-message">✅ ประมวลผลสำเร็จ {success_count} ไฟล์</p>', unsafe_allow_html=True)
            if error_count > 0:
                st.markdown(f'<p class="error-message">❌ ประมวลผลไม่สำเร็จ {error_count} ไฟล์</p>', unsafe_allow_html=True)
                
            status.update(label="✅ เสร็จสิ้น!", state="complete")
else:
    st.info("กรุณาอัปโหลดไฟล์ KML อย่างน้อย 1 ไฟล์")
